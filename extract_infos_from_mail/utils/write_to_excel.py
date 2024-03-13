from openpyxl import Workbook
from extract_infos_from_mail.utils.constants.HEADERS import HEADERS


def write_to_excel(excel_data, file_path):
    wb = Workbook()
    ws = wb.active

    column = 1
    for header in HEADERS:
        title = header["title"]
        columns = header["columns"]

        # Écrire le titre
        ws.cell(row=1, column=column, value=title)

        # Fusionner les cellules pour le titre
        ws.merge_cells(
            start_row=1,
            start_column=column,
            end_row=1,
            end_column=column + len(columns) - 1,
        )

        # Écrire les noms des colonnes
        for i, column_name in enumerate(columns):
            ws.cell(row=2, column=column + i, value=column_name)

        # Mettre à jour la colonne pour le prochain titre
        column += len(columns)

    for row in excel_data:
        ws.append(row)

    # Sauvegarder le fichier
    wb.save(file_path)
